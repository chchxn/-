"""
excel_parser.py
층별일람표.xlsx에서 대오더 정보와 화상검사 포인트를 파싱합니다.
"""

import openpyxl
from dataclasses import dataclass, field
from typing import Optional


@dataclass
class OrderInfo:
    """대오더 바코드 라벨 시트의 한 행"""
    no: int
    barcode_text: str
    order_info: str       # 예: 3029C003AA
    serial_prefix: str    # 예: 2EQ02001


@dataclass
class InspectionPoint:
    """화상검사Point 시트의 한 부품 행"""
    part_no: str           # 구분 번호 (예: C550, 1525)
    part_name: str         # 부품명
    quantity: int
    part_code: str         # 부품 번호 (예: FE2-C550-000)
    guarantee: str         # 보증 항목
    # 대오더별 체크포인트: { order_info: "①" or "●" or "⑤, ⑦" or None }
    check_points: dict = field(default_factory=dict)


def parse_orders(xlsx_path: str) -> list[OrderInfo]:
    """대오더 바코드 라벨 시트 파싱"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['대오더 바코드 라벨']
    orders = []
    for row in ws.iter_rows(min_row=3, values_only=True):  # 3행부터 데이터
        no, barcode, _, order, serial = row
        if no is None:
            continue
        orders.append(OrderInfo(
            no=int(no),
            barcode_text=str(barcode) if barcode else "",
            order_info=str(order) if order else "",
            serial_prefix=str(serial) if serial else "",
        ))
    return orders


def parse_inspection_points(xlsx_path: str) -> tuple[list[str], list[InspectionPoint]]:
    """
    화상검사Point 시트 파싱
    Returns:
        order_columns: 열 순서대로 정렬된 대오더 정보 목록
        points: 검사 포인트 목록
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['화상검사Point']

    # 5행: 대오더 정보 헤더 (열 9~25)
    order_row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    order_columns = []  # (col_index, order_info)
    for col_idx in range(9, ws.max_column + 1):
        val = order_row[col_idx - 1]
        if val and str(val).startswith('30') or (val and str(val).startswith('31')):
            order_columns.append((col_idx, str(val)))

    # 10행부터 데이터 행
    points = []
    for row in ws.iter_rows(min_row=10, max_row=ws.max_row, values_only=True):
        part_no = row[1]   # B열
        part_name = row[2] # C열
        quantity = row[4]  # E열
        part_code = row[5] # F열
        guarantee = row[6] # G열

        if part_no is None:
            continue

        check_points = {}
        for col_idx, order_info in order_columns:
            val = row[col_idx - 1]
            if val is not None:
                check_points[order_info] = str(val)

        points.append(InspectionPoint(
            part_no=str(part_no),
            part_name=str(part_name) if part_name else "",
            quantity=int(quantity) if quantity else 0,
            part_code=str(part_code) if part_code else "",
            guarantee=str(guarantee) if guarantee else "",
            check_points=check_points,
        ))

    return [o for _, o in order_columns], points


def get_steps_for_order(order_info: str, points: list[InspectionPoint]) -> dict[str, str]:
    """
    특정 대오더에 해당하는 검사 포인트와 STEP 번호 반환
    Returns: { part_no: step_str }  예: { "C550": "①", "1525": "⑤, ⑦" }
    """
    result = {}
    for pt in points:
        cp = pt.check_points.get(order_info)
        if cp and cp != '●':   # ●는 생산이력 없음 표시 (일단 제외)
            result[pt.part_no] = cp
    return result


# 원형 문자 → 정수 변환 테이블
CIRCLE_NUM_MAP = {
    '①': 1, '②': 2, '③': 3, '④': 4, '⑤': 5,
    '⑥': 6, '⑦': 7, '⑧': 8, '⑨': 9, '⑩': 10,
    '⑪': 11, '⑫': 12, '⑬': 13, '⑭': 14, '⑮': 15,
    '⑯': 16, '⑰': 17, '⑱': 18, '⑲': 19, '⑳': 20,
}


def parse_step_numbers(step_str: str) -> list[int]:
    """'⑤, ⑦' → [5, 7]"""
    steps = []
    for ch in step_str:
        if ch in CIRCLE_NUM_MAP:
            steps.append(CIRCLE_NUM_MAP[ch])
    return steps


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "층별일람표.xlsx"
    
    orders = parse_orders(path)
    print(f"대오더 수: {len(orders)}")
    for o in orders[:3]:
        print(f"  NO{o.no}: {o.order_info} / {o.serial_prefix} / {o.barcode_text[:20]}...")
    
    order_cols, points = parse_inspection_points(path)
    print(f"\n검사 포인트 수: {len(points)}")
    print(f"대오더 열 수: {len(order_cols)}")
    
    print(f"\n첫 번째 대오더 {order_cols[0]} 검사 포인트:")
    steps = get_steps_for_order(order_cols[0], points)
    for part, step in steps.items():
        nums = parse_step_numbers(step)
        print(f"  {part}: {step} → step 파일 번호 {nums}")


# ── 바코드 이미지 추출 ────────────────────────────────────────

def extract_master_images(xlsx_path: str, data_dir: str):
    """
    엑셀 '대오더 바코드 라벨' 시트의 바코드 이미지를
    각 대오더 폴더에 master.png 로 저장합니다.

    data/
    └── 3029C003AA/
        └── master.png   ← 여기 저장
    """
    import zipfile as zf
    import openpyxl
    import os

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['대오더 바코드 라벨']

    # 행 번호 → 대오더 정보 (3행부터 데이터, NO 1부터)
    row_to_order = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        no, _, _, order, _ = row
        if no is not None:
            row_to_order[int(no) + 2] = str(order)

    saved = []
    with zf.ZipFile(xlsx_path) as z:
        for i, (row_no, order_info) in enumerate(row_to_order.items(), start=1):
            img_src  = f'xl/media/image{i}.png'
            out_dir  = os.path.join(data_dir, order_info)
            out_path = os.path.join(out_dir, 'master.png')

            if os.path.exists(out_path):
                continue  # 이미 있으면 스킵

            os.makedirs(out_dir, exist_ok=True)
            try:
                with z.open(img_src) as src:
                    with open(out_path, 'wb') as dst:
                        dst.write(src.read())
                saved.append(order_info)
            except Exception:
                pass  # 이미지 없는 대오더는 스킵

    if saved:
        print(f"[마스터 이미지] {len(saved)}개 저장 완료")
    return saved
